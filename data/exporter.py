"""Excel export using openpyxl.  Three sheets: Lap Data, Summary, Fuel Analysis."""
from __future__ import annotations
import datetime
from typing import Optional

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from telemetry.packet import format_laptime_display
from telemetry.state import LapRecord


_HEADER_FONT  = Font(bold=True, color="FFFFFF")
_HEADER_FILL  = PatternFill("solid", fgColor="1F4E78")
_ALT_FILL     = PatternFill("solid", fgColor="DCE6F1")
_CENTER       = Alignment(horizontal="center")


def export_to_excel(
    records: list[LapRecord],
    filepath: str,
    session_label: str = "",
) -> None:
    """Write records to an xlsx file with three sheets."""
    wb = Workbook()

    _build_lap_sheet(wb.active, records, session_label)
    wb.active.title = "Lap Data"

    _build_summary_sheet(wb.create_sheet("Summary"), records, session_label)
    _build_fuel_sheet(wb.create_sheet("Fuel Analysis"), records)

    wb.save(filepath)


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------

_PRACTICE_FILL = PatternFill("solid", fgColor="DCE6F1")  # light blue — default alt row
_RACE_FILL     = PatternFill("solid", fgColor="D6EAD6")  # light green
_QUAL_FILL     = PatternFill("solid", fgColor="D6D6EA")  # light purple
_PIT_FILL      = PatternFill("solid", fgColor="FFF2CC")  # yellow


def _build_lap_sheet(ws, records: list[LapRecord], session_label: str) -> None:
    headers = [
        "Lap", "Session", "Lap Time", "Lap Time (ms)", "Delta (s)",
        "Best Lap", "Fuel Start (L)", "Fuel End (L)", "Fuel Used (L)",
        "Avg Fuel/Lap (L)", "Position", "Pit Stop", "Timestamp",
    ]
    _write_headers(ws, headers)

    running_fuel_total = 0.0
    running_count = 0

    for i, r in enumerate(records, start=2):
        running_fuel_total += r.fuel_used
        running_count += 1
        avg_fuel = running_fuel_total / running_count if running_count else 0.0

        stype = r.session_type.value.capitalize() if r.session_type else "Practice"

        ws.cell(i, 1,  r.lap_num)
        ws.cell(i, 2,  stype)
        ws.cell(i, 3,  format_laptime_display(r.lap_time_ms))
        ws.cell(i, 4,  r.lap_time_ms)
        ws.cell(i, 5,  round(r.delta_ms / 1000, 3) if r.best_lap_ms > 0 else "")
        ws.cell(i, 6,  format_laptime_display(r.best_lap_ms))
        ws.cell(i, 7,  round(r.fuel_start, 2))
        ws.cell(i, 8,  round(r.fuel_end, 2))
        ws.cell(i, 9,  round(r.fuel_used, 2))
        ws.cell(i, 10, round(avg_fuel, 2))
        ws.cell(i, 11, r.position if r.position > 0 else "")
        ws.cell(i, 12, "Yes" if r.is_pit_lap else "")
        ws.cell(i, 13, r.timestamp)

        # Row fill: pit stop overrides session colour; even rows get session tint
        if r.is_pit_lap:
            row_fill = _PIT_FILL
        elif stype == "Race":
            row_fill = _RACE_FILL if i % 2 == 0 else None
        elif stype == "Qualifying":
            row_fill = _QUAL_FILL if i % 2 == 0 else None
        else:
            row_fill = _PRACTICE_FILL if i % 2 == 0 else None

        if row_fill:
            for col in range(1, len(headers) + 1):
                ws.cell(i, col).fill = row_fill

    _auto_width(ws)
    ws.freeze_panes = "A2"


def _build_summary_sheet(ws, records: list[LapRecord], session_label: str) -> None:
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 20

    rows = []
    if session_label:
        rows.append(("Session", session_label))
    rows.append(("Export Time", datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")))
    rows.append(("Total Laps", len(records)))

    valid_times = [r.lap_time_ms for r in records if r.lap_time_ms > 0]
    rows.append(("Best Lap Time", format_laptime_display(min(valid_times)) if valid_times else "--"))
    avg_ms = int(sum(valid_times) / len(valid_times)) if valid_times else 0
    rows.append(("Average Lap Time", format_laptime_display(avg_ms)))

    total_fuel = sum(r.fuel_used for r in records)
    rows.append(("Total Fuel Used (L)", round(total_fuel, 2)))
    fuel_vals = [r.fuel_used for r in records if r.fuel_used > 0]
    avg_fuel = sum(fuel_vals) / len(fuel_vals) if fuel_vals else 0.0
    rows.append(("Average Fuel/Lap (L)", round(avg_fuel, 2)))

    pit_stops = sum(1 for r in records if r.is_pit_lap)
    rows.append(("Pit Stops", pit_stops))

    for row_idx, (key, val) in enumerate(rows, start=1):
        ws.cell(row_idx, 1, key).font = Font(bold=True)
        ws.cell(row_idx, 2, val)


def _build_fuel_sheet(ws, records: list[LapRecord]) -> None:
    headers = [
        "Lap", "Fuel Used (L)", "Cumulative Fuel (L)",
        "Projected Total (L)", "Pit Stop",
    ]
    _write_headers(ws, headers)

    cumulative = 0.0
    running_fuel: list[float] = []

    for i, r in enumerate(records, start=2):
        cumulative += r.fuel_used
        if r.fuel_used > 0:
            running_fuel.append(r.fuel_used)
        total_laps = len(records)
        avg_fuel = sum(running_fuel) / len(running_fuel) if running_fuel else 0.0
        projected = avg_fuel * total_laps

        ws.cell(i, 1, r.lap_num)
        ws.cell(i, 2, round(r.fuel_used, 2))
        ws.cell(i, 3, round(cumulative, 2))
        ws.cell(i, 4, round(projected, 2) if projected > 0 else "")
        ws.cell(i, 5, "Yes" if r.is_pit_lap else "")

        if i % 2 == 0:
            for col in range(1, len(headers) + 1):
                ws.cell(i, col).fill = _ALT_FILL

    _auto_width(ws)
    ws.freeze_panes = "A2"


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _write_headers(ws, headers: list[str]) -> None:
    for col, title in enumerate(headers, start=1):
        cell = ws.cell(1, col, title)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _CENTER


def _auto_width(ws) -> None:
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value or "")))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)
